import openpyxl
import os

database_path = "scouting_database"
sheet_path = "test.xlsx"

locations = {
    "match_num":              ["c3" ,"c3" ,"c3" ,"c3" ], 

    "atop":                   ["d21","d51","d51","l10"],
    "amiddle":                ["d22","d52","d52","m10"],
    "abottom":                ["d23","d53","d53","n10"],
    "aengaged":               ["k11","k11","k11","d10"],
    "adocked":                ["k9" ,"k9" ,"k9" ,"d7" ],
    "aattempted":             [""   ,""   ,"l9" ,"d13"],
    "mobility":               ["i12","i12","i12","d16"],

    "ttop":                   ["d31","d61","d61","l12"],
    "tmiddle":                ["d32","d62","d62","m12"],
    "tbottom":                ["d33","d63","d63","n12"],
    "tengaged":               ["k17","k17","k17","h10"],
    "tdocked":                ["k15","k15","k15","h7" ],
    "tattempted":             [""   ,""   ,"l15","h13"],
    "parked":                 ["i18","i18","i18","h16"],

    "malfunction":            ["h5" ,"g5" ,"g5" ,"i12"],
    "defense":                ["g5" ,"f5" ,"f5" ,"i15"],
    "scout_name":             [""   ,"a3" ,"a3" ,"a3" ],
    "intake_notes":           [""   ,"a22","a22","c23"],
    "driving_notes":          [""   ,"a25","a25","c26"],
    "notable_plays_or_fouls": [""   ,"a28","a28","c29"],
}

alpha = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

def read_folder(team_num: str) -> list:
    folder_path = f"{database_path}\\{team_num}"
    if not os.path.isdir(folder_path): return f"Folder not found: {folder_path}"

    folder = os.listdir(folder_path)

    return list(filter(folder_filter, folder))

def folder_filter(file_name: str) -> bool:
    if file_name == "Pits": return False

    if file_name[0] == "~": return False

    return True

def filter_database(folder: str) -> bool:
    try:
        int(folder)
        return True
    except:
        return False

def get_version_id(read) -> int:
    # in the locations dict 
    # each version of the sheet has different ids
    # because not all the sheets are exactly the same
    if read["A1"].value == "V2" or read["A2"].value == "V2": 
        return 1
    if read["A1"].value == "V3": 
        return 2
    if read["A1"].value == "V4":
        return 3
    return 0

# returns the id of row
def get_last_updated_row(master) -> str:
    found = False
    i = 3
    while not found:
        # print(master[f"A{i}"].value)
        if master[f"A{i}"].value == None: break
        i+=1
    return i

def get_all_filenames(master) -> list:
    row = 3
    names = []
    while True:
        if not master[f"A{row}"].value: break
        names.append(master[f"X{row}"].value)

        row+=1

    return names

# gets date of a match from file_name
def parse_file_name(file_name: str) -> str:
    start_id = file_name.rfind("(")
    end_id = file_name.rfind(")")
    
    return file_name[start_id+1:end_id]

def open_sheet(num, file_name):
    wb = openpyxl.load_workbook(f"{database_path}\\{num}\\{file_name}", read_only=True, data_only=True)
    ws = wb.active
    return ws

# read is the sheet to be read from 
# and master is, well, the master
def update_row(row: str, read, master, team_num: str, file_name):
    master[f"A{row}"] = team_num
    master[f"B{row}"] = parse_file_name(file_name)

    v_id = get_version_id(read)

    for id, val in enumerate(locations.values()):
        letter = alpha[id + 2]
        key = val[v_id]

        if key == "": continue
        # print(read[key].value)
        master[f"{letter}{row}"] = read[key].value
    
    master[f"{alpha[id+3]}{row}"] = file_name

def update_error(row: str, master, file_name):
    master[f"A{row}"] = f"ERROR: {file_name} could not be opened"

def main():
    # master_wb = openpyxl.Workbook()
    try:
        master_wb = openpyxl.load_workbook(sheet_path)
        master_ws = master_wb["Database"]
    except:
        master_wb = openpyxl.Workbook()
        master_ws = master_wb.active

    team_nums = filter(filter_database,os.listdir(database_path))
    # team_nums = ["25"]

    logged_files = get_all_filenames(master_ws)

    row = get_last_updated_row(master_ws)
    for team_num in team_nums:    
        files = read_folder(team_num)
        # print(team_num)
        for file in files:
            if file in logged_files: continue
            # print(file)
            try:
                read = open_sheet(team_num, file)
                update_row(f"{row}", read, master_ws, team_num, file)
            except:
                update_error(f"{row}", master_ws, file)
            row+=1

    master_wb.save(sheet_path)
    os.system(f'start EXCEL.EXE "{sheet_path}"')

main()