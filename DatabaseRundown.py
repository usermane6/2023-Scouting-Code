import os
from openpyxl import load_workbook, Workbook

def load_all(path:str) -> list:
    if not os.path.isdir(path): return

    paths = os.listdir(path)
    paths.remove("Pits")

    print(paths)

    all_sheets = []

    for p in paths:
        try:
            all_sheets.append(load_workbook(f"{path}\\{p}", data_only=True).active)
        except:
            continue

    return all_sheets

def get_rundown(database_path:str, path:str, num:str) -> None:
    sheets = load_all(path)

    request = {
        "mobility":     [],
        "adocked":      [],
        "aengaged":     [],
        "parked":       [],
        "tdocked":      [],
        "tengaged":     [],
        "atop":         [],
        "amiddle":      [],
        "abottom":      [],
        "ttop":         [],
        "tmiddle":      [],
        "tbottom":      [],
        "atotalpieces": [],
        "ttotalpieces": [],
        "malfunction":  [],
        "defense":      [],
        "aaccuracy":    [],
        "taccuracy":    [],
    }

    locations = {
        "mobility":     ["i12","i12","i12"],
        "adocked":      ["k9" ,"k9" ,"k9" ],
        "aengaged":     ["k11","k11","k11"],
        "parked":       ["i18","i18","i18"],
        "tdocked":      ["k15","k15","k15"],
        "tengaged":     ["k17","k17","k17"],
        "atop":         ["d21","d51","d51"],
        "amiddle":      ["d22","d52","d52"],
        "abottom":      ["d23","d53","d53"],
        "ttop":         ["d31","d61","d61"],
        "tmiddle":      ["d32","d62","d62"],
        "tbottom":      ["d33","d63","d63"],
        "atotalpieces": ["d24","d54","d54"],
        "ttotalpieces": ["d34","d64","d64"],
        "malfunction":  ["h5" ,"g5" ,"g5" ],
        "defense":      ["g5" ,"f5" ,"f5" ],
        "aaccuracy":    [""   ,""   ,"l9" ],
        "taccuracy":    [""   ,""   ,"l15"],
    }

    notes = {
        "intake_notes": [],
        "driving_notes": [],
        "notable_plays_or_fouls": [],
    }

    note_locations = {
        "intake_notes":           ["", "A22", "A22"],
        "driving_notes":          ["", "A25", "A25"],
        "notable_plays_or_fouls": ["", "A28", "A28"],
    }

    v3_aeng_sum = 0 # sum of total auton engage 
    v3_teng_sum = 0 # sum of total teleop engage

    v3s = []

    for sheet in sheets:
        if sheet["B3"].value == None: continue
        id = 0
        # print(sheet["A1"].value)
        if sheet["A1"].value == "V2" or sheet["A2"].value == "V2": 
            id = 1
        if sheet["A1"].value == "V3": 
            id = 2
            v3s.append(sheet)
            # print("V3 here")

        for key in request.keys():

            if id == 2 and key == "aengaged" and sheet[locations[key][id]].value:
                v3_aeng_sum += 1
            if id == 2 and key == "tengaged" and sheet[locations[key][id]].value:
                v3_teng_sum += 1

            if id != 2 and key in ("aaccuracy", "taccuracy"): continue

            request[key].append(sheet[locations[key][id]].value)
        
        # print(notes.keys())
        for key in notes.keys():
            if id == 0: break

            notes[key].append(sheet[note_locations[key][id]].value)

    request_workbook = Workbook() 
    rs = request_workbook.active # request sheet

    rs["b1"] = "avg/percent"
    rs["c1"] = "max"
    rs["d1"] = "min"

    for i, key in enumerate(request.keys()):
        # print(key, request[key])
        rs[f"a{i+2}"] = key
        # rs[f"b{i+2}"] = avg(request[key])

        if not request[key]: continue

        if key == "aaccuracy":
            print("v3_aeng_sum", v3_aeng_sum)
            # value is the total engages over the total attempts
            # this gets the accuracy of the engaging
            try:
                rs[f"b{i+2}"] = percent(v3_aeng_sum / sum(request[key]))
            except:
                rs[f"b{i+2}"] = 0
            continue

        if key == "taccuracy":
            print("v3_teng_sum", v3_teng_sum)
            try:
                rs[f"b{i+2}"] = percent(v3_teng_sum / sum(request[key]))
            except:
                rs[f"b{i+2}"] = 0
            continue

        if type(request[key][0]) == bool: 
            rs[f"b{i+2}"] = percent(avg(request[key]))
            continue

        rs[f"b{i+2}"] = round(avg(request[key]), 2)
        rs[f"c{i+2}"] = max(request[key])
        rs[f"d{i+2}"] = min(request[key])

    ids=["f", "g", "h"]
    for i, key in enumerate(notes.keys()):
        # print(i, key)
        current_id = ids[i]
        rs[f"{current_id}1"] = key
        # print(notes[key])
        j = 2
        for note in notes[key]:
            if note == None: continue
            rs[f"{current_id}{j}"] = note
            j+=1

        
    workbook_path = f"{database_path}\\Rundowns\\RUNDOWN{num}.xlsx"
    request_workbook.save(workbook_path)
    os.system(f'start EXCEL.EXE "{workbook_path}"')

def avg(vals) -> float: 
    return sum(vals) / len(vals)

def percent(val: float) -> str:
    percent = round(val * 100, 2)
    return f"{percent}%"
